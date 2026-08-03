"""Stage the 'Fourni par Citizing / Donnees Distribution.xlsm' export into
CSVs shaped for import into the Repousse schema (taxa, users,
planting_projects, distribution events/records).

Output is intermediate staging data, not a 1:1 dump of the final Ecto
tables: some columns (project management granularity, distribution slot
dates) need product decisions before a real migration script is written.
See the printed report for open questions.

Usage: python3 scripts/data-prep/prepare_distributions.py
"""

import csv
import difflib
import re
import unicodedata
from collections import Counter
from pathlib import Path

import openpyxl

EMAIL_RE = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]+$")

ROOT = Path(__file__).resolve().parents[2]
SOURCE = ROOT / "data" / "Fourni par Citizing" / "Données Distribution.xlsm"
OUT = ROOT / "data" / "prepared"

CATEGORY_COLS = ["Pti fruitiers", "Grimpantes", "Arbustes", "Fruitiers", "Arbres", "Inconnu"]

# Typos/abbreviations seen in the "Especes" column that don't match the
# Typologie reference by exact or fuzzy string match. Built by manual
# inspection of the 158 raw values against the 132-name Typologie sheet.
MANUAL_SPECIES_FIXES = {
    "aconia": "Aronia",
    "aromia": "Aronia",
    "herberis": "Berberis",
    "naquis": "Maquis",
    "maquis du chili": "Maquis",
    "licyet": "Goji - Lyciet",
    "lyciet": "Goji - Lyciet",
    "lyciets": "Goji - Lyciet",
    "lysiets": "Goji - Lyciet",
    "lissiers": "Goji - Lyciet",
    "lisier": "Goji - Lyciet",
    "goji": "Goji - Lyciet",
    "cotonesaster": "Cotoneaster",
    "cotonesater": "Cotoneaster",
    "cotonester": "Cotoneaster",
    "cotonaester": "Cotoneaster",
    "costonaer": "Cotoneaster",
    "cotoneasters": "Cotoneaster",
    "sichuan": "Poivrier sichuan",
    "timut": "Poivrier timut",
    "raisinier de chine": "Raisinier",
    "botte de saule": "Saule",
    "hetre": "Hêtre",
    "indetermine": "Inconnu",
    "inconnu": "Inconnu",
    "orme champetre erreur ? confusion avec erable champetre": "Orme champêtre",
    "cotoneaster": "Cotoneaster",
    "consoude": "Consoude",
}
# Species that occur in the source but have no matching entry at all in
# the Typologie reference sheet (132 canonical names vs 158 raw labels) —
# added to taxa.csv with a blank category, pending the fuller species
# reference the client is preparing separately.
MISSING_FROM_TYPOLOGIE = {"Cotoneaster", "Consoude"}
# Raw values with no reasonable species match: junk fragments or
# free-text notes left in the source cell rather than a species name.
UNRESOLVED_SPECIES = {"environ", "maquereaux", "maquis"}


def normalize(s):
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    return re.sub(r"\s+", " ", s)


def load_typologie(wb):
    ws = wb["Typologie"]
    taxa = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        if not name:
            continue
        category = next((CATEGORY_COLS[i] for i in range(6) if row[i + 1] == "X"), None)
        taxa.append({"name": name.strip(), "category": category, "nb_arbre": row[7]})
    return taxa


def build_species_lookup(taxa):
    by_norm = {normalize(t["name"]): t["name"] for t in taxa}
    canonical_names = list(by_norm.values())

    def resolve(raw):
        if not raw:
            return None, None, "empty"
        norm = normalize(raw)
        if norm in by_norm:
            return by_norm[norm], "exact", 1.0
        if norm in MANUAL_SPECIES_FIXES:
            fixed = MANUAL_SPECIES_FIXES[norm]
            return fixed, "manual_fix", 1.0
        if norm in UNRESOLVED_SPECIES:
            return None, "unresolved", 0.0
        match = difflib.get_close_matches(raw.strip(), canonical_names, n=1, cutoff=0.72)
        if match:
            return match[0], "fuzzy", difflib.SequenceMatcher(None, norm, normalize(match[0])).ratio()
        return None, "no_match", 0.0

    return resolve


def clean_email(raw):
    if not raw:
        return None
    value = str(raw).strip().lower()
    if EMAIL_RE.match(value):
        return value
    # common AZERTY typo: "à" (shift+0) typed instead of "@" (altgr+0)
    swapped = value.replace("à", "@")
    if EMAIL_RE.match(swapped):
        return swapped
    return None


def normalize_phone(value):
    if not value:
        return ""
    digits = re.sub(r"\D", "", str(value))
    return digits[-9:] if digits else ""


def summarize_phone(phones):
    """Picks a representative phone number and a confidence level.

    Some campaigns have a copy-paste/fill-handle bug where the phone
    increments by 1 on every extra species line for the same person
    (789912516, 789912517, 789912518...) instead of repeating the real
    number — seen on ~70 beneficiaries, not just one or two. Detected as 3+
    normalized values forming a consecutive run.

    Confidence:
      high   - no broken-sequence pattern detected
      medium - broken sequence, but one value repeats (>=2) -> likely real
      low    - broken sequence, all values distinct -> no way to tell which
               one (if any) is real; the pick is an arbitrary placeholder
    """
    if not phones:
        return "", "high"
    norm_digits = sorted({int(normalize_phone(p)) for p in phones if normalize_phone(p).isdigit()})
    is_broken = len(norm_digits) >= 3 and all(b - a == 1 for a, b in zip(norm_digits, norm_digits[1:]))
    if not is_broken:
        return phones[0], "high"

    counts = Counter(normalize_phone(p) for p in phones)
    best_norm, best_count = counts.most_common(1)[0]
    representative = next(p for p in phones if normalize_phone(p) == best_norm)
    return representative, ("medium" if best_count >= 2 else "low")


def as_int(value):
    if isinstance(value, (int, float)):
        return int(value)
    return None


def load_distribution_rows(wb):
    ws = wb["Données distributions"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        rows.append(
            {
                "annee": row[0],
                "beneficiaire_type": row[1],
                "code_distribution": row[2],
                "cp": row[3],
                "departement": row[4],
                "ville": row[5],
                "nom_prenom": (row[6] or "").strip(),
                "email": clean_email(row[7]),
                "telephone": row[8],
                "racines_nues": as_int(row[9]),
                "nombre": as_int(row[10]),
                "espece_raw": row[11],
                "typologie": row[12],
            }
        )
    return rows


def split_name(full_name, beneficiaire_type):
    parts = full_name.split()
    if beneficiaire_type == "Particulier" and len(parts) == 2:
        return parts[0], parts[1], False
    return full_name, "", True


def write_csv(path, fieldnames, records):
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(records)
    print(f"wrote {len(records):5d} rows -> {path.relative_to(ROOT)}")


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(SOURCE, data_only=True, read_only=True)

    taxa = load_typologie(wb)
    resolve_species = build_species_lookup(taxa)
    dist_rows = load_distribution_rows(wb)

    # ---- taxon_categories.csv + taxa.csv ----
    categories = sorted({t["category"] for t in taxa if t["category"]})
    write_csv(
        OUT / "taxon_categories.csv",
        ["slug", "name"],
        [{"slug": normalize(c).replace(" ", "-"), "name": c} for c in categories],
    )
    taxa_rows = [
        {
            "common_name": t["name"],
            "category_slug": normalize(t["category"]).replace(" ", "-") if t["category"] else "",
            "is_non_taxonomic": True,
            "source_nb_arbre": t["nb_arbre"],
        }
        for t in taxa
    ]
    taxa_rows += [
        {"common_name": name, "category_slug": "", "is_non_taxonomic": True, "source_nb_arbre": ""}
        for name in sorted(MISSING_FROM_TYPOLOGIE)
    ]
    write_csv(
        OUT / "taxa.csv",
        ["common_name", "category_slug", "is_non_taxonomic", "source_nb_arbre"],
        taxa_rows,
    )

    # ---- species_mapping_review.csv ----
    raw_species = {}
    for r in dist_rows:
        if r["espece_raw"]:
            raw_species.setdefault(r["espece_raw"], 0)
            raw_species[r["espece_raw"]] += 1
    mapping_rows = []
    for raw, count in sorted(raw_species.items(), key=lambda kv: -kv[1]):
        matched, method, score = resolve_species(raw)
        mapping_rows.append(
            {
                "raw_value": raw,
                "occurrences": count,
                "matched_taxon": matched or "",
                "method": method,
                "confidence": round(score, 2) if isinstance(score, float) else score,
                "needs_review": method in ("fuzzy", "no_match", "unresolved"),
            }
        )
    write_csv(
        OUT / "species_mapping_review.csv",
        ["raw_value", "occurrences", "matched_taxon", "method", "confidence", "needs_review"],
        mapping_rows,
    )

    # ---- beneficiaries: split rows with/without email ----
    with_email = [r for r in dist_rows if r["email"]]
    without_email = [r for r in dist_rows if not r["email"]]

    # ---- users.csv + planting_projects.csv (one row per unique email) ----
    by_email = {}
    for r in with_email:
        by_email.setdefault(r["email"], []).append(r)

    users, projects = [], []
    for idx, (email, records) in enumerate(sorted(by_email.items()), start=1):
        first = records[0]
        temp_id = f"U{idx:04d}"
        first_name, last_name, needs_review = split_name(first["nom_prenom"], first["beneficiaire_type"])
        phones = [str(r["telephone"]) for r in records if r["telephone"]]
        phone, phone_confidence = summarize_phone(phones)
        users.append(
            {
                "temp_id": temp_id,
                "email": email,
                "first_name": first_name,
                "last_name": last_name,
                "full_name_raw": first["nom_prenom"],
                "phone": phone,
                "telephone_confidence": phone_confidence,
                "hanko_id": "",
                "status": "active",
                "role": "member",
                "adhesion_active": False,
                "needs_name_review": needs_review,
            }
        )

        types = {r["beneficiaire_type"] for r in records if r["beneficiaire_type"]}
        villes = {r["ville"] for r in records if r["ville"]}
        cps = {r["cp"] for r in records if r["cp"]}
        depts = {r["departement"] for r in records if r["departement"]}
        beneficiaire_type = next(iter(types), "")
        projects.append(
            {
                "temp_id": f"P{idx:04d}",
                "owner_temp_id": temp_id,
                "name": f"{first['nom_prenom']} ({first['ville'] or 'ville inconnue'})",
                "management_type": "individual" if beneficiaire_type == "Particulier" else "collective",
                "beneficiaire_type_raw": beneficiaire_type,
                "address": next(iter(villes), ""),
                "postal_code": next(iter(cps), ""),
                "department": next(iter(depts), ""),
            }
        )

    write_csv(
        OUT / "users.csv",
        [
            "temp_id", "email", "first_name", "last_name", "full_name_raw",
            "phone", "telephone_confidence", "hanko_id", "status", "role",
            "adhesion_active", "needs_name_review",
        ],
        users,
    )
    write_csv(
        OUT / "planting_projects.csv",
        [
            "temp_id", "owner_temp_id", "name", "management_type",
            "beneficiaire_type_raw", "address", "postal_code", "department",
        ],
        projects,
    )

    # ---- distribution_events.csv (one row per campaign code) ----
    events = {}
    for r in dist_rows:
        code = r["code_distribution"] or "(sans code)"
        e = events.setdefault(code, {"code": code, "annees": set(), "n_records": 0, "total_qty": 0})
        e["annees"].add(str(r["annee"]))
        e["n_records"] += 1
        e["total_qty"] += r["nombre"] or 0
    sorted_events = sorted(events.values(), key=lambda e: -e["n_records"])
    write_csv(
        OUT / "distribution_events.csv",
        ["temp_id", "code", "annees", "n_records", "total_qty", "status"],
        [
            {
                "temp_id": f"E{i + 1:04d}",
                **e,
                "annees": "|".join(sorted(e["annees"])),
                "status": "closed",
            }
            for i, e in enumerate(sorted_events)
        ],
    )

    # ---- recover no-email rows that share a phone or exact name with a
    # known user (their email is on a different distribution line) ----
    phone_owners, name_owners = {}, {}
    for u in users:
        # skip "low" confidence phones: an arbitrary pick among an all-distinct
        # broken sequence, matching on it would be a coincidence, not a signal
        if u["telephone_confidence"] != "low":
            pn = normalize_phone(u["phone"])
            if pn:
                phone_owners.setdefault(pn, set()).add(u["temp_id"])
        nn = normalize(u["full_name_raw"])
        name_owners.setdefault(nn, set()).add(u["temp_id"])
    phone_owners = {p: next(iter(ids)) for p, ids in phone_owners.items() if len(ids) == 1}
    name_owners = {n: next(iter(ids)) for n, ids in name_owners.items() if len(ids) == 1}

    recovered, unrecoverable = [], []
    for r in without_email:
        pn = normalize_phone(r["telephone"])
        nn = normalize(r["nom_prenom"])
        temp_id = phone_owners.get(pn) or name_owners.get(nn)
        if temp_id:
            recovered.append((r, temp_id, "telephone" if pn in phone_owners else "nom_exact"))
        else:
            unrecoverable.append(r)

    # ---- no_email_beneficiaries.csv: one row per remaining person, not
    # per distribution line, so the association has an actual contact list ----
    by_person = {}
    for r in unrecoverable:
        key = normalize(r["nom_prenom"])
        person = by_person.setdefault(key, {"rows": [], "phones": set()})
        person["rows"].append(r)
        if r["telephone"]:
            person["phones"].add(str(r["telephone"]))

    no_email_out = []
    for key, group in sorted(by_person.items()):
        rows = group["rows"]
        first = rows[0]
        villes = {r["ville"] for r in rows if r["ville"]}
        especes = [f"{r['espece_raw'] or r['typologie']} x{r['nombre'] or '?'}" for r in rows]
        _, phone_confidence = summarize_phone(list(group["phones"]))
        no_email_out.append(
            {
                "nom_prenom": first["nom_prenom"],
                "telephones": "|".join(sorted(group["phones"])),
                "telephone_confidence": phone_confidence,
                "ville": "|".join(sorted(villes)),
                "n_lignes": len(rows),
                "especes_qty": "; ".join(especes),
            }
        )
    write_csv(
        OUT / "no_email_beneficiaries.csv",
        ["nom_prenom", "telephones", "telephone_confidence", "ville", "n_lignes", "especes_qty"],
        no_email_out,
    )

    # ---- distribution_records.csv (flat staging, one row per source line) ----
    email_to_temp = {u["email"]: u["temp_id"] for u in users}
    records_out = []
    for r in with_email:
        matched, method, score = resolve_species(r["espece_raw"])
        records_out.append(
            {
                "user_temp_id": email_to_temp[r["email"]],
                "annee": r["annee"],
                "code_distribution": r["code_distribution"] or "",
                "typologie": r["typologie"],
                "espece_raw": r["espece_raw"] or "",
                "espece_matched": matched or "",
                "racines_nues": r["racines_nues"] or "",
                "nombre": r["nombre"] or "",
            }
        )
    for r, temp_id, _method in recovered:
        matched, method, score = resolve_species(r["espece_raw"])
        records_out.append(
            {
                "user_temp_id": temp_id,
                "annee": r["annee"],
                "code_distribution": r["code_distribution"] or "",
                "typologie": r["typologie"],
                "espece_raw": r["espece_raw"] or "",
                "espece_matched": matched or "",
                "racines_nues": r["racines_nues"] or "",
                "nombre": r["nombre"] or "",
            }
        )
    write_csv(
        OUT / "distribution_records.csv",
        [
            "user_temp_id", "annee", "code_distribution", "typologie",
            "espece_raw", "espece_matched", "racines_nues", "nombre",
        ],
        records_out,
    )

    # ---- report ----
    review_count = sum(1 for m in mapping_rows if m["needs_review"])
    name_review_count = sum(1 for u in users if u["needs_name_review"])
    users_medium = sum(1 for u in users if u["telephone_confidence"] == "medium")
    users_low = sum(1 for u in users if u["telephone_confidence"] == "low")
    print()
    print("=== rapport ===")
    print(f"taxa reference        : {len(taxa)}")
    print(f"especes brutes        : {len(raw_species)} ({review_count} a valider)")
    print(f"utilisateurs (email)  : {len(users)} ({name_review_count} noms a verifier)")
    print(f"telephones fiabilite  : {users_medium} moyenne (majorite trouvee), {users_low} basse (pick arbitraire)")
    print(f"lignes sans email     : {len(without_email)} dont {len(recovered)} recuperees (tel/nom deja connu)")
    print(f"personnes a contacter  : {len(no_email_out)}")
    print(f"projets               : {len(projects)}")
    print(f"campagnes (evenements): {len(events)}")
    print(f"lignes distribution   : {len(records_out)}")


if __name__ == "__main__":
    main()
